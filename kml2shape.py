import sqlite3
from openpyxl import load_workbook


conn = sqlite3.connect('1.db')
cur = conn.cursor()
cur.execute('DROP TABLE IF EXISTS routes')
cur.execute("create table routes (ruta, abscisa)")
wb = load_workbook(filename='Matriz de Distancia/Matriz Distancia SAE1.xlsx',
                   read_only=True)
ws = wb['Hoja1']
for i, row in enumerate(ws):
    if i is 0:
        continue
    ruta = row[0].value
    abscisa = row[5].value
    cur.execute("insert into routes values (?, ?)", (ruta, abscisa))
conn.commit()
cur.close()
conn.close()


conn = sqlite3.connect('2.db')
cur = conn.cursor()
cur.execute('DROP TABLE IF EXISTS routes')
cur.execute("create table routes (linea, ruta, abscisa)")
wb = load_workbook(filename='Matriz de Distancia/Matriz Distancia SAE2.xlsx',
                   read_only=True)
ws = wb['Hoja1']
for i, row in enumerate(ws):
    if i is 0:
        continue
    linea = row[0].value
    ruta = row[1].value
    abscisa = row[3].value
    cur.execute("insert into routes values (?, ?, ?)", (linea, ruta, abscisa))
conn.commit()
cur.close()
conn.close()